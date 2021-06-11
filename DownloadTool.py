# coding=utf-8
import wx
import requests
import threading
import os
from urllib.request import urlretrieve
import urllib
# from urllib import parse
from openpyxl import load_workbook
import time
# import pandas as pd
from bs4 import BeautifulSoup
import re
import socket
import datetime

# 设置超时时间为30s
socket.setdefaulttimeout(30)

APP_TITLE = u'下载文件工具'
APP_ICON = 'res/python.ico'


class mainFrame(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((460, 420))
        self.Center()

        self.selectExcelPathBtn = wx.Button(self, -1, u'选择文件夹', pos=(10, 20), size=(100, -1), style=wx.ALIGN_LEFT)

        # wx.StaticText(self, -1, u'请求地址：', pos=(10, 20), size=(60, -1), style=wx.ALIGN_LEFT)
        self.excelFile = wx.TextCtrl(self, -1, '', pos=(130, 20), size=(260, -1), name='excelFile', style=wx.TE_LEFT)

        # self.selectOutPathBtn = wx.Button(self, -1, u'下载文件路径', pos=(10, 50), size=(100, -1), style=wx.ALIGN_LEFT)
        # self.outPath = wx.TextCtrl(self, -1, '', pos=(130, 50), size=(260, -1), name='outPath', style=wx.TE_LEFT)

        wx.StaticText(self, -1, u'Cookie：', pos=(10, 50), size=(60, -1), style=wx.ALIGN_LEFT)
        self.cookie = wx.TextCtrl(self, -1, 'SESSION=bd5f9323-30a0-481d-98ad-18697fdd3f24', pos=(130, 50),
                                  size=(260, -1), name='Cookie', style=wx.TE_LEFT)

        wx.StaticText(self, -1, u'下载时段：', pos=(10, 80), size=(60, -1), style=wx.ALIGN_LEFT)
        self.startHour = wx.TextCtrl(self, -1, '20', pos=(130, 80), size=(40, -1), name='开始时', style=wx.TE_LEFT)
        self.endHour = wx.TextCtrl(self, -1, '8', pos=(180, 80), size=(40, -1), name='结束时', style=wx.TE_LEFT)


        self.btn_start = wx.Button(self, -1, u'开始下载', pos=(10, 130), size=(80, -1))

        self.area = wx.TextCtrl(self, -1, '', pos=(10, 170), size=(380, 200), name='area',
                                style=wx.TE_LEFT | wx.TE_MULTILINE)

        self.Bind(wx.EVT_BUTTON, self.OnSelectExcel, self.selectExcelPathBtn)
        # self.Bind(wx.EVT_BUTTON, self.OnSelectOutPath, self.selectOutPathBtn)
        self.btn_start.Bind(wx.EVT_LEFT_DOWN, self.startWork)

        # 设置是否暂停
        self.pause = False

    def OnSelectExcel(self, event):
        dlg = wx.DirDialog(self, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            print(dlg.GetPath())  # 文件夹路径
            self.excelFile.SetLabelText(dlg.GetPath())
        dlg.Destroy()

    # def OnSelectOutPath(self, event):
    #     dlg = wx.DirDialog(self, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
    #     if dlg.ShowModal() == wx.ID_OK:
    #         print(dlg.GetPath())  # 文件夹路径
    #         self.outPath.SetLabelText(dlg.GetPath())
    #     dlg.Destroy()

    def startWork(self, event):
        if self.excelFile.GetValue() == '':
            self.area.AppendText("请选择Excel文件\n")
            return
        if self.cookie.GetValue() == '':
            self.area.AppendText("请输入cookie\n")
            return
            # if self.year.GetValue() == '':
            #     self.area.AppendText("请输入年份\n")
            return

        t1 = threading.Thread(target=self.pre_work)
        t1.start()
        t2 = threading.Thread(target=self.keep_alive)
        t2.setDaemon(True)
        t2.start()

    def pre_work(self):

        if self.timeInDate() == False:
            self.pause = True

        self.excelFile.Disable()
        self.selectExcelPathBtn.Disable()
        # self.outPath.Disable()
        self.cookie.Disable()
        self.btn_start.Disable()
        # self.year.Disable()
        self.area.AppendText("开始查找Excel文件\n")
        fileDict = self.find_excel()
        if len(fileDict.items()) <= 0:
            self.area.AppendText("未找到Excel文件\n")
            self.excelFile.Enable()
            self.cookie.Enable()
            self.btn_start.Enable()
            self.selectExcelPathBtn.Enable()
            return

        for file, year in fileDict.items():
            self.area.Clear()
            self.area.AppendText("开始加载：" + file + "\n")
            ent_list = self.read_excel(file)
            t1 = time.time()
            size = len(ent_list)
            if size <= 0:
                self.area.AppendText("加载企业失败,结束下载\n")
                continue

            self.area.AppendText("加载完毕，开始下载文件\n")
            for r in range(0, size):
                while True:
                    if self.pause == False:
                        break
                    else:
                        self.area.AppendText("当前时间未在指定时间段内，等待中...\n")
                        time.sleep(30)

                entInfo = ent_list[r]
                ent_name = entInfo.get("entName")
                hangye = entInfo.get("hangye")
                print(ent_name)
                if os.path.exists(os.path.join(self.excelFile.GetValue(), year, hangye)) == False:
                    os.makedirs(os.path.join(self.excelFile.GetValue(), year, hangye))
                self.download(ent_name, hangye, year)
                time.sleep(0.3)
            t2 = time.time()
            t3 = int(t2) - int(t1)
            self.area.AppendText("下载耗时秒：" + str(t3) + "\n")

        self.area.AppendText("全部下载结束\n")
        self.excelFile.Enable()
        # self.outPath.Enable()
        self.cookie.Enable()
        self.btn_start.Enable()
        self.selectExcelPathBtn.Enable()
        # self.year.Enable()

    def keep_alive(self):
        print("刷新系统保持活跃")
        header = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': 'http://10.100.248.214/manager/reportInfo/list?hasReport=false',
            'x-requested-with': 'XMLHttpRequest',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Accept-Encoding': 'gzip, deflate',
            'Host': '10.100.248.214',
            'Origin': 'http://10.100.248.214',
            'Pragma': 'no-cache',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
            'Cookie': self.cookie.GetValue()
        }
        while True:
            if self.timeInDate() == False:
                self.pause = True
            # 调用系统接口，保持session活跃
            ret = requests.get('http://10.100.248.214/manager/index',  timeout=10, headers=header)
            time.sleep(30)

    def timeInDate(self):
        curtHour = datetime.datetime.now().hour
        starthour = self.startHour.GetValue()
        endhour = self.endHour.GetValue()
        if starthour != "" and endhour != "":
            sh = int(starthour)
            eh = int(endhour)
            if eh < sh and curtHour>=sh and curtHour<24:
                return True
            if eh > sh and curtHour>=sh and curtHour<eh:
                return True
        else:
            return True

        return False



    def find_excel(self):
        dict = {}
        for root, dirs, files in os.walk(self.excelFile.GetValue()):
            print(root)  # 当前目录路径
            print(dirs)  # 当前路径下所有子目录
            print(files)  # 当前路径下所有非目录子文件
            for file in files:
                fileNameArr = os.path.splitext(file)
                if fileNameArr[1] == '.xlsx' or fileNameArr[1] == '.xls':
                    fileInfoArr = fileNameArr[0].split("-")
                    if len(fileInfoArr) > 2:
                        self.area.AppendText("找到文件：" + file + "\n")
                        dict[file] = fileInfoArr[2]  # os.path.join(self.excelFile.GetValue(), file)
                        # 创建年份文件夹
                        if os.path.exists(self.excelFile.GetValue() + "\\" + fileInfoArr[2]) == False:
                            os.makedirs(self.excelFile.GetValue() + "\\" + fileInfoArr[2])
        return dict

    def read_excel(self, file):
        ent_list = []
        print('excel：', os.path.join(self.excelFile.GetValue(), file))
        wb = load_workbook(os.path.join(self.excelFile.GetValue(), file))
        sheet = wb.active
        rnum = sheet.max_row + 1
        cnum = sheet.max_column
        for r in range(3, rnum):
            ent_name = sheet.cell(row=r, column=2).value
            hangye = sheet.cell(row=r, column=3).value
            # print(ent_name)
            if ent_name is None:
                continue
            entInfo = {"entName": ent_name, "hangye": hangye}
            ent_list.append(entInfo)
        return ent_list

    def download(self, ent_name, hangye, year):
        # down_url = "http://www.baidu.com/" + parse.quote(ent_name) + ".pdf"
        # self.download_file2(down_url, self.outPath.GetLabel())
        self.searchEnt(ent_name, hangye, year)
        return True

    def download_file1(self, url, store_path):
        filename = url.split("/")[-1]
        filepath = os.path.join(store_path, filename)

        file_data = requests.get(url, allow_redirects=True).content
        with open(filepath, 'wb') as handler:
            handler.write(file_data)

    def download_file2(self, url, filename, store_path):
        while True:
            try:
                # store_path = store_path.encode(encoding='UTF-8',errors='strict')
                # filename = url.split("/")[-1]
                filepath = os.path.join(store_path, filename)

                # def callbackfunc(blocknum, blocksize, totalsize):
                #     percent = 100.0 * blocknum * blocksize / totalsize
                #     if percent > 100:
                #         percent = 100
                #     self.area.AppendText(filename + "下载进度:" + str(percent) + "%\n")
                self.area.AppendText("[" + filename + "]文件下载中...\n")
                urlretrieve(url, filepath)
                self.area.AppendText("[" + filename + "]文件下载完毕...\n")
                break
            except socket.timeout:
                count = 1
                while count <= 5:
                    try:
                        urllib.request.urlretrieve(url, filepath)
                        break
                    except socket.timeout:
                        err_info = "[" + filename + ']下载超时,重新发起下载请求第%d次' % count
                        print(err_info)
                        self.area.AppendText(err_info + '\n')
                        count += 1
                if count > 5:
                    print("downloading picture fialed!")
            except urllib.error.HTTPError as e:
                self.area.AppendText("[" + filename + "]下载异常：" + e.reason + '\n')
                print(e.reason)
                print(e.code)
                time.sleep(3)

    def searchEnt(self, ent_name, hangye, year):
        header = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': 'http://10.100.248.214/manager/reportInfo/list?hasReport=false',
            'x-requested-with': 'XMLHttpRequest',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Accept-Encoding': 'gzip, deflate',
            'Host': '10.100.248.214',
            'Origin': 'http://10.100.248.214',
            'Pragma': 'no-cache',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
            'Cookie': self.cookie.GetValue()
        }
        paramData = {
            'industryName': '',
            'industryCode': '',
            'provinceCode': '',
            'cityCode': '',
            'countyCode': '',
            'permitCode': '',
            'companyName': ent_name,
            'startFztime': '',
            'endFztime': '',
            'fzjg': False,
            'first': '',
            'isUpdate': '',
            'page': 1
        }
        try:
            print("获取企业信息列表")
            ret = requests.post('http://10.100.248.214/manager/reportInfo/list?hasReport=true', data=paramData,
                                timeout=30, headers=header)
            # df = pd.read_html(ret.text)[0]   # 该页面只返回一个表格，所以取第0个表格
            # print("获取企业HTML:@@@@@@@@@", ret.text, "&&&&&&&&&")
            soup = BeautifulSoup(ret.text, 'html.parser')
            a_ctx = soup.findAll("a", {'target': 'rightFrame'})  # 抓取a标签
            req_url2 = ''
            for ax in a_ctx:
                req_url2 = ax.get('href')
                print('获取到企业详情url', req_url2)

            if len(req_url2) <= 0:
                print('没有查询到企业')
                self.area.AppendText('没有查询到企业[' + ent_name + "]信息\n")
                return
            req_url2 = 'http://10.100.248.214' + req_url2
            self.downloadPdf(ent_name, year, hangye, req_url2)

        except requests.exceptions.ConnectionError:
            self.area.AppendText('连接服务器超时\n')
            print('[ERROR]ConnectionError -- will retry connect')

    def downloadPdf(self, ent_name, year, hangye, req_url2):
        header = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': 'http://10.100.248.214/manager/reportInfo/list?hasReport=false',
            'x-requested-with': 'XMLHttpRequest',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Accept-Encoding': 'gzip, deflate',
            'Host': '10.100.248.214',
            'Origin': 'http://10.100.248.214',
            'Pragma': 'no-cache',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
            'Cookie': self.cookie.GetValue()
        }

        req_url3 = 'http://10.100.248.214/report/pdf/month?reportId=dataid&token=a247c4ab8ab04ff1b470c9ff3bdd95fb'
        try:
            # 获取2018年的数据
            print("年份：", year)
            ret = requests.get(req_url2 + '&year=' + year, timeout=30, headers=header)

            # print('获取到数据HTML:@@@@@@@:', ret.text, "&&&&&&&&")
            soup = BeautifulSoup(ret.text, 'html.parser')
            s_ctx = soup.findAll("span", {'class': 'page-count'})  # 抓取总条数
            a_ctx = soup.findAll("a", {'class': 'btn-base btn-noborder icon-download'})  # 抓取a标签 获取数据id
            totalCount = int(s_ctx[0].contents[1])
            totalPage = int(totalCount / 10) + 1

            findFlag = False
            for ax in a_ctx:
                parent_text = ax.parent.parent.text
                if parent_text.find("年报") == -1:
                    continue

                data_herf = ax.get('href')
                if data_herf == '':
                    continue
                data_id = re.findall("\d+", data_herf)[0]
                print('获取到数据id:', data_id)
                findFlag = True

                # 在循环内一个个开始下载文件
                req_url4 = req_url3.replace("dataid", data_id)
                filePath = os.path.join(self.excelFile.GetValue(), year, hangye)
                self.download_file2(req_url4, ent_name + ".pdf", filePath)

            if findFlag == False and totalPage > 1:
                time.sleep(0.3)
                for pageNo in (2, totalPage):
                    ret = requests.get(req_url2 + '&year=' + year + '&pageNo=' + str(pageNo), timeout=30,
                                       headers=header)
                    soup = BeautifulSoup(ret.text, 'html.parser')
                    a_ctx = soup.findAll("a", {'class': 'btn-base btn-noborder icon-download'})
                    for ax in a_ctx:
                        parent_text = ax.parent.parent.text
                        if parent_text.find("年报") == -1:
                            continue
                        data_herf = ax.get('href')
                        if data_herf == '':
                            continue
                        data_id = re.findall("\d+", data_herf)[0]
                        print('获取到数据id:', data_id)
                        findFlag = True

                        # 在循环内一个个开始下载文件
                        req_url4 = req_url3.replace("dataid", data_id)
                        filePath = os.path.join(self.excelFile.GetValue(), year, hangye)
                        self.download_file2(req_url4, ent_name + ".pdf", filePath)
                    if findFlag == True:
                        break

                    time.sleep(0.3)

            if findFlag == False:
                self.area.AppendText('企业[' + ent_name + "]" + year + "没有年报，跳过下载\n")

        except requests.exceptions.ConnectionError:
            print('[ERROR]ConnectionError -- will retry connect')
        except Exception as ex:
            print("下载遇到遇到异常")
            print(ex)


class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True


if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()
