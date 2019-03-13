import wx
import wx.lib.filebrowsebutton
from openpyxl import load_workbook
import json,os

APP_TITLE = u'表格处理'
APP_ICON = 'res/python.ico'

class mainFrame(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((520, 300))
        self.Center()

        self.btn_selectFile = wx.lib.filebrowsebutton.FileBrowseButton(self,pos=(20, 40),size=(400,-1),style=wx.ALIGN_RIGHT,labelText='', fileMask='*.xlsx',buttonText='打开',changeCallback=self.OnRead)

        self.tip = wx.StaticText(self, -1, u'', pos=(20, 100), size=(400, -1), style=wx.ST_NO_AUTORESIZE)

        self.obj1={}
        self.obj2={}
        self.obj1['@@@']=[]


    def On_size(self, evt):
        # 改变窗口大小事件函数
        self.Refresh()
        evt.Skip()

    def OnRead(self, event):
        """"""
        self.btn_selectFile.Disable()
        self.tip.SetLabel('数据处理中，请不要关闭')

        f = open(os.path.abspath('.')+"\\config.json", encoding='utf-8')
        if f is None:
            self.tip.SetLabel('配置文件config.json不存在，请放在同级目录')
            return
        conf={}
        try:
           conf = json.load(f)
        except:
            self.tip.SetLabel('config.json不是json格式，请检查')
            return

        for key in conf:
            self.obj1[conf[key]]=[]
            self.obj2[conf[key]]=[]

        print(self.btn_selectFile.GetValue())
        wb = load_workbook(self.btn_selectFile.GetValue())
        sheet = wb.active
        rnum=sheet.max_row
        cnum=sheet.max_column
        for r in range(3,rnum):
            value1=sheet.cell(row=r, column=3).value
            value2=sheet.cell(row=r, column=5).value
            value3=sheet.cell(row=r, column=6).value
            # print(value1,value2,value3)
            if value1 is None and value2 is None:
                continue
            if value2 is None:
                continue
            if value1 is None :
                value1=" @@@= "
            if value3 is not None:
                value3=value3.replace(' ', '')
            a1=value1.split("=")
            atype=a1[0].split(" ")[1]
            #print(atype,',',a1[1],',',value2,',',value3)
            # if atype == 'A':
            #     atype='X'
            # if atype == 'BB':
            #     atype='YY'
            # if atype == 'CC':
            #     atype='OOO'
            # if atype == 'DX':
            #     atype='UUUU'
            if atype in conf:
                atype = conf[atype]
            obj1={'mc':0,'type':atype, 'value':a1[1],'prefix':a1[0], 'row':r}  # mc--匹配度  type--类型  value--字符串 row--所在字符串的行数 @占位
            obj2={'type':value3, 'value':value2, 'row':r}
            self.obj1[atype].append(obj1)
            self.obj2[value3].append(obj2)

        print('---------------------------------------------')
        for key in self.obj2:
            arr1=self.obj1[key]
            arr2=self.obj2[key]

            for obj2 in arr2:
                for obj1 in arr1:
                    value = obj1['value'].replace('-', '').replace('315', '').lower()  # 处理-，处理315特殊干扰，处理英文匹配
                    tarr = self.cut(value)
                    for t in tarr:
                        v2=obj2['value'].replace('315', '').lower()
                        if v2.find(t) != -1:
                            obj1['mc'] += 1
                # 查找完毕，开始排序,默认最大放在最前面
                arr1.sort(key=lambda obj: obj['mc'], reverse=True)
                #print(arr1[0],obj2['value'])
                if len(arr1)>0 and arr1[0]['mc'] > 0:
                    sheet.cell(row=obj2['row'], column=4).value = arr1[0]['prefix']+'='+arr1[0]['value']
                for ob in arr1:
                    ob['mc']=0

        wb.save(self.btn_selectFile.GetValue().split(".xls")[0]+'_new.xlsx')
        self.btn_selectFile.Enable()
        self.tip.SetLabel('数据处理完成')


    # 找出字符串所有子串
    def cut(self,s):
        results = []
        # x + 1 表示子字符串长度
        for x in range(len(s)):
            # i 表示偏移量
            for i in range(len(s) - x):
                results.append(s[i:i + x + 1])
        return results



class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True

if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()
