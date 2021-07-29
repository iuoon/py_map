# coding=utf-8
import wx
import threading
import os
import openpyxl
from openpyxl import load_workbook
import time
from bs4 import BeautifulSoup
import requests
import shutil
import pandas as pd
import numpy as np

APP_TITLE = u'工具'
APP_ICON = 'res/python.ico'


class mainFrame(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((460, 420))
        self.Center()

        self.selectFilePathBtn = wx.Button(self, -1, u'输出文件路径', pos=(10, 20), size=(100, -1), style=wx.ALIGN_LEFT)
        self.filePath = wx.TextCtrl(self, -1, '', pos=(130, 20), size=(260, -1), name='filePath', style=wx.TE_LEFT)

        self.btn_start = wx.Button(self, -1, u'开始', pos=(10, 130), size=(80, -1))

        self.area = wx.TextCtrl(self, -1, '', pos=(10, 170), size=(380, 200), name='area',
                                style=wx.TE_LEFT | wx.TE_MULTILINE)

        self.Bind(wx.EVT_BUTTON, self.OnSelectFilePath, self.selectFilePathBtn)
        self.btn_start.Bind(wx.EVT_LEFT_DOWN, self.startWork)

        # 设置是否暂停
        self.pause = False

    def OnSelectFilePath(self, event):
        dlg = wx.DirDialog(self, u"选择输出文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            print(dlg.GetPath())  # 文件夹路径
            self.filePath.SetLabelText(dlg.GetPath())
        dlg.Destroy()

    def startWork(self, event):
        if self.filePath.GetValue() == '':
            self.area.AppendText("请选择输出文件夹\n")
            return
        t1 = threading.Thread(target=self.pre_work)
        t1.start()

    def pre_work(self):
        self.filePath.Disable()
        self.selectFilePathBtn.Disable()
        self.btn_start.Disable()

        self.area.AppendText("获取行业列表数据\n")
        self.GetIndustryDataList('', self.filePath.GetLabelText())
        self.area.AppendText("处理完成\n")
        self.filePath.Enable()
        self.selectFilePathBtn.Enable()
        self.btn_start.Enable()

    ## 按行业获取数据列表
    def GetIndustryDataList(self, industryName, filePath):
        header = {
            'Host': 'permit.mee.gov.cn',
            'Connection': 'keep-alive',
            'Origin': 'http://permit.mee.gov.cn',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36',
            'Accept': "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            'Referer': 'http://permit.mee.gov.cn/perxxgkinfo/syssb/xkgg/xkgg!licenseInformation.action',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cookie': 'JSESSIONID=EEF9F68EF33E259F42AEA1A6ED6EEC7C'
        }
        paramData = {
            'page.pageNo': 1,
            'page.orderBy': '',
            'page.order': '',
            'province': '',
            'city': '',
            'registerentername': '',
            'xkznum': '',
            'treadname': '电力、热力生产和供应业,电力生产,火力发电,热电联产,水力发电,核力发电,风力发电,太阳能发电,生物质能发电,生物质能发电-生活垃圾焚烧发电,其他电力生产,电力供应,热力生产和供应',
            'treadcode': 'D44,D441,D4411,D4412,D4413,D4414,D4415,D4416,D4417,D4417-1,D4419,D442,D443',
            'publishtime': ''
        }
        try:
            print("获取企业信息列表")
            ret = requests.post('http://permit.mee.gov.cn/perxxgkinfo/syssb/xkgg/xkgg!licenseInformation.action',
                                data=paramData,
                                timeout=30, headers=header)
            # print("获取列表数据HTML: ", ret.text)
            df = pd.read_html(ret.text)[0]  # 该页面只返回一个表格，所以取第0个表格
            print(df.values)
            soup = BeautifulSoup(ret.text, 'html.parser')
            href_ctx = soup.findAll("td", {'class': 'bgcolor1'})

            # df.to_excel(filePath + "\\result.xlsx", index=False, header=False)
            detailPath = filePath + "\\result.xlsx"
            if os.path.exists(detailPath) == True:
                os.remove(detailPath)
            wb0 = openpyxl.Workbook()
            sheet = wb0.active
            cnum = sheet.max_column
            rows_old = sheet.max_row
            dataLen = len(df.values)
            for i in range(0, dataLen):
                detail_url = ''
                if i > 0:
                    a = href_ctx[i - 1].next
                    detail_url = 'http://permit.mee.gov.cn' + a.get('href')
                data = []
                for j in range(0, len(df.values[i])):
                    if j < len(df.values[i]) - 1:
                        data.append(df.values[i][j])
                if detail_url != '':
                    data.append(detail_url)
                sheet.append(data)
            wb0.save(detailPath)

            total_ctx = soup.findAll("div", {'class': 'fr margin-t-33 margin-b-20'})[0]  # 抓取总条数
            total_text = total_ctx.text
            total_text = total_text.split("\r\n")[2]
            total_text = total_text.replace('\t', '')
            total_text = total_text.replace(' ', '')
            total_text = total_text.replace('共', '')
            total_text = total_text.replace('页', '')
            total = int(total_text)
            totalPage = 1
            if total % 10 == 0:
                totalPage = int(total / 10)
            else:
                totalPage = int(total / 10) + 1
            if totalPage < 2:
                return
            for pageNo in range(2, totalPage):
                time.sleep(0.3)
                paramData['page.pageNo'] = pageNo
                ret = requests.post('http://permit.mee.gov.cn/perxxgkinfo/syssb/xkgg/xkgg!licenseInformation.action',
                                    data=paramData,
                                    timeout=30, headers=header)
                df2 = pd.read_html(ret.text)[0]  # 该页面只返回一个表格，所以取第0个表格
                print(df2.values)
                soup = BeautifulSoup(ret.text, 'html.parser')
                href_ctx = soup.findAll("td", {'class': 'bgcolor1'})

                wb = load_workbook(detailPath)
                sheet = wb.active
                cnum = sheet.max_column
                rows_old = sheet.max_row
                dataLen = len(df2.values)
                for i in range(1, dataLen):
                    detail_url = ''
                    if i > 0:
                        a = href_ctx[i - 1].next
                        detail_url = 'http://permit.mee.gov.cn' + a.get('href')
                    data = []
                    for j in range(0, len(df.values[i])):
                        if j < len(df.values[i]) - 1:
                            data.append(df.values[i][j])
                    if detail_url != '':
                        data.append(detail_url)
                    sheet.append(data)
                wb.save(detailPath)







        except requests.exceptions.ConnectionError:
            self.area.AppendText('连接服务器超时\n')
            print('[ERROR]ConnectionError -- will retry connect')

    def searchEnt(self, ent_name, hangye, year):
        header = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Referer': 'http://10.100.248.214/manager/reportInfo/list?hasReport=false',
            'x-requested-with': 'XMLHttpRequest',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN',
            'Cache-Control': 'max-age=0',
            'Connection': 'keep-alive',
            'Accept-Encoding': 'gzip, deflate',
            'Host': '10.100.248.214',
            'Origin': 'http://10.100.248.214',
            'Pragma': 'no-cache',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
            'Cookie': self.cookie.GetValue()
        }
        paramData = {
            'industryName': '',
            'industryCode': '',
            'provinceCode': '',
            'cityCode': '',
            'countyCode': '',
            'permitCode': '',
            'companyName': '',
            'startFztime': '',
            'endFztime': '',
            'fzjg': False,
            'first': '',
            'isUpdate': '',
            'page': 1
        }
        try:
            print("获取企业信息列表")
            ret = requests.post('http://10.100.248.214/manager/reportInfo/list?hasReport=true', data=paramData,
                                timeout=30, headers=header)
            # df = pd.read_html(ret.text)[0]   # 该页面只返回一个表格，所以取第0个表格
            # print("获取企业HTML:@@@@@@@@@", ret.text, "&&&&&&&&&")
            soup = BeautifulSoup(ret.text, 'html.parser')
            a_ctx = soup.findAll("a", {'target': 'rightFrame'})  # 抓取a标签
            req_url2 = ''
            for ax in a_ctx:
                req_url2 = ax.get('href')
                print('获取到企业详情url', req_url2)

            if len(req_url2) <= 0:
                print('没有查询到企业')
                self.area.AppendText('没有查询到企业[' + ent_name + "]信息\n")
                return
            req_url2 = 'http://10.100.248.214' + req_url2
            self.downloadPdf(ent_name, year, hangye, req_url2)

        except requests.exceptions.ConnectionError:
            self.area.AppendText('连接服务器超时\n')
            print('[ERROR]ConnectionError -- will retry connect')


class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True


if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()
