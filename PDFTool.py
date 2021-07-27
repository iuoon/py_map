# coding=utf-8
import wx
import threading
import os
from openpyxl import load_workbook
import time
from bs4 import BeautifulSoup
import shutil

APP_TITLE = u'工具'
APP_ICON = 'res/python.ico'


class mainFrame(wx.Frame):

    def __init__(self, parent):
        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((460, 420))
        self.Center()

        self.selectExcelPathBtn = wx.Button(self, -1, u'选择Excel文件', pos=(10, 20), size=(100, -1), style=wx.ALIGN_LEFT)

        # wx.StaticText(self, -1, u'请求地址：', pos=(10, 20), size=(60, -1), style=wx.ALIGN_LEFT)
        self.excelFile = wx.TextCtrl(self, -1, '', pos=(130, 20), size=(260, -1), name='excelFile', style=wx.TE_LEFT)

        self.selectPdfPathBtn = wx.Button(self, -1, u'PDF文件路径', pos=(10, 50), size=(100, -1), style=wx.ALIGN_LEFT)
        self.pdfPath = wx.TextCtrl(self, -1, '', pos=(130, 50), size=(260, -1), name='pdfPath', style=wx.TE_LEFT)

        self.btn_start = wx.Button(self, -1, u'开始', pos=(10, 130), size=(80, -1))

        self.area = wx.TextCtrl(self, -1, '', pos=(10, 170), size=(380, 200), name='area',
                                style=wx.TE_LEFT | wx.TE_MULTILINE)

        self.Bind(wx.EVT_BUTTON, self.OnSelectExcel, self.selectExcelPathBtn)
        self.Bind(wx.EVT_BUTTON, self.OnSelectPDFPath, self.selectPdfPathBtn)
        self.btn_start.Bind(wx.EVT_LEFT_DOWN, self.startWork)

        # 设置是否暂停
        self.pause = False

    def OnSelectExcel(self, event):
        dlg = wx.FileDialog(self, u"选择Excel文件", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            print(dlg.GetPath())  # 文件夹路径
            self.excelFile.SetLabelText(dlg.GetPath())
        dlg.Destroy()

    def OnSelectPDFPath(self, event):
        dlg = wx.DirDialog(self, u"选择PDF文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            print(dlg.GetPath())  # 文件夹路径
            self.pdfPath.SetLabelText(dlg.GetPath())
        dlg.Destroy()

    def startWork(self, event):
        if self.excelFile.GetValue() == '':
            self.area.AppendText("请选择Excel文件\n")
            return
        if self.excelFile.GetValue() == '':
            self.area.AppendText("请选择Excel文件\n")
            return

        t1 = threading.Thread(target=self.pre_work)
        t1.start()

    def pre_work(self):
        self.excelFile.Disable()
        self.selectExcelPathBtn.Disable()
        self.selectPdfPathBtn.Disable()
        self.pdfPath.Disable()
        self.btn_start.Disable()

        self.area.AppendText("加载：" + self.excelFile.GetValue() + "\n")
        t1 = time.time()
        wb = load_workbook(self.excelFile.GetValue())
        sheet = wb.active
        rnum = sheet.max_row + 1
        cnum = sheet.max_column
        for r in range(2, rnum):
            ent_name = sheet.cell(row=r, column=1).value
            # print(ent_name)
            if ent_name is None:
                continue
            print(ent_name)
            self.area.AppendText("开始解析：" + ent_name + ".pdf\n")
            pdpath = os.path.join(self.pdfPath.GetValue(), ent_name + ".pdf")
            if os.path.exists(pdpath) == False:
                self.area.AppendText("未找到：" + ent_name + ".pdf，跳过\n")
                continue
            if os.path.exists(os.path.join(self.pdfPath.GetValue(), "temp")) == False:
                os.makedirs(os.path.join(self.pdfPath.GetValue(), "temp"))
            else:
                shutil.rmtree(os.path.join(self.pdfPath.GetValue(), "temp"))
                os.makedirs(os.path.join(self.pdfPath.GetValue(), "temp"))
            ret = os.system(
                "pdf2htmlEX\\pdf2htmlEX.exe -f 3 -l 6 --zoom 1 --dest-dir " + os.path.join(self.pdfPath.GetValue(),
                                                                                           "temp") + " " + pdpath)
            print(ret)
            fileName = ""
            for root, dirs, files in os.walk(os.path.join(self.pdfPath.GetValue(), "temp")):
                if len(files) <= 0:
                    self.area.AppendText("文件已损坏：" + ent_name + ".pdf，跳过\n")
                    continue
                fileName = files[0]
                break
            if fileName == "":
                continue
            with open(os.path.join(self.pdfPath.GetValue(), "temp", fileName), "r", encoding="UTF-8") as f:  # 打开文件
                data = f.read()  # 读取文件
                soup = BeautifulSoup(data, 'html.parser')
                div_list = soup.find_all('div')
                for i in range(0, len(div_list)):
                    ctx = div_list[i].text.strip(' ')
                    if ctx.startswith("熟料产量"):
                        # print(ctx)
                        text_sl = ctx.replace("熟料产量", "").replace(" ", "").replace("吨", "")
                        # print(text_sl)
                        sheet.cell(row=r, column=3).value = text_sl
                    if ctx.startswith("耗煤量"):
                        # print(ctx)
                        text_hm = ctx.replace("耗煤量", "").replace(" ", "").replace("吨", "")
                        # print(text_hm)
                        sheet.cell(row=r, column=4).value = text_hm
            self.area.AppendText(ent_name + ".pdf 解析完毕\n")
        self.area.AppendText(self.excelFile.GetValue() + " 更新\n")
        wb.save(self.excelFile.GetValue())

        t2 = time.time()
        t3 = int(t2) - int(t1)
        self.area.AppendText("耗时：" + str(t3) + "秒\n")

        self.area.AppendText("处理完成\n")
        self.excelFile.Enable()
        self.selectPdfPathBtn.Enable()
        self.selectExcelPathBtn.Enable()
        self.pdfPath.Enable()
        self.btn_start.Enable()

    def read_excel(self, file):
        ent_list = []
        print('excel：', os.path.join(self.excelFile.GetValue(), file))
        wb = load_workbook(os.path.join(self.excelFile.GetValue(), file))
        sheet = wb.active
        rnum = sheet.max_row + 1
        cnum = sheet.max_column
        for r in range(3, rnum):
            ent_name = sheet.cell(row=r, column=2).value
            hangye = sheet.cell(row=r, column=3).value
            # print(ent_name)
            if ent_name is None:
                continue
            entInfo = {"entName": ent_name, "hangye": hangye}
            ent_list.append(entInfo)
        return ent_list


class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True


if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()
