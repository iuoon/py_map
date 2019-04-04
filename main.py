#!/usr/bin/python
# -*- coding: UTF-8 -*-
from string import digits

import wx
import requests
import time
import os
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger
import threading
import city
from openpyxl import Workbook,load_workbook


APP_TITLE = u'爬取数据-爬限速正式版'
APP_ICON = 'res/python.ico'

class mainFrame(wx.Frame):
    '''程序主窗口类，继承自wx.Frame'''

    def __init__(self, parent):
        '''构造函数'''

        wx.Frame.__init__(self, parent, -1, APP_TITLE)
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.SetSize((600, 480))
        self.Center()

        # if hasattr(sys, "frozen") and getattr(sys, "frozen") == "windows_exe":
        #     exeName = win32api.GetModuleFileName(win32api.GetModuleHandle(None))
        #     icon = wx.Icon(exeName, wx.BITMAP_TYPE_ICO)
        # else :
        #     icon = wx.Icon(APP_ICON, wx.BITMAP_TYPE_ICO)
        # self.SetIcon(icon)

        wx.StaticText(self, -1, u'设置高德地图Key：', pos=(10, 20), size=(100, -1), style=wx.ALIGN_RIGHT)
        # tip
        self.tip = wx.StaticText(self, -1, u'爬限速每天爬一次，点开始会立即执行爬一次，注意操作，以免封禁', pos=(20, 340), size=(400, -1), style=wx.ST_NO_AUTORESIZE)
        self.tip2 = wx.StaticText(self, -1, u"矩形坐标全用英文逗号隔开，格式为：左下角经纬度,右上角经纬度 ", pos=(20, 360), size=(400, -1), style=wx.ST_NO_AUTORESIZE)
        self.tip3 = wx.StaticText(self, -1, u"取经纬度坐标地址：https://lbs.amap.com/console/show/picker", pos=(20, 380), size=(400, -1), style=wx.ST_NO_AUTORESIZE)
        self.tip3 = wx.StaticText(self, -1, u"若设置了经纬度矩形区域，则开始时按矩形抓取，否则按选择城市抓取", pos=(20, 400), size=(400, -1), style=wx.ST_NO_AUTORESIZE)
        # key
        self.gd_key = wx.TextCtrl(self, -1, '', pos=(130, 20), size=(200, -1), name='GD_KEY', style=wx.TE_LEFT)

        self.area = wx.TextCtrl(self, -1, '', pos=(10, 60), size=(320, 200), name='area', style=wx.TE_LEFT | wx.TE_MULTILINE)



        self.btn_start = wx.Button(self, -1, u'开始', pos=(350, 20), size=(100, 25))
        self.btn_pause = wx.Button(self, -1, u'暂停', pos=(350, 50), size=(100, 25))

        #self.btn_close = wx.Button(self, -1, u'关闭窗口', pos=(350, 80), size=(100, 25))
        wx.StaticText(self, -1, u'设置矩形坐标点：', pos=(350, 80), size=(100, -1), style=wx.ALIGN_LEFT)
        self.loat = wx.TextCtrl(self, -1, '', pos=(350, 100), size=(200, -1), name='loat', style=wx.TE_LEFT)


        wx.StaticText(self, -1, u'选择省市：', pos=(350, 150), size=(50, -1), style=wx.ALIGN_LEFT)
        pros=[]
        for key in city.province_city:
            pros.append(key)
        self.ch1 = wx.ComboBox(self,-1,value='选择省',choices=pros,pos=(350, 170))
        self.ch2 = wx.ComboBox(self,-1,value='选择市',choices=[],pos=(350, 200))
        self.city = '-1'  # 当前城市
        self.preCity = '-1'  # 上一次城市
        self.cityTip = wx.StaticText(self, -1, u'当前城市：未选择', pos=(350, 230), size=(400, -1), style=wx.ST_NO_AUTORESIZE)

        self.btn_start2 = wx.Button(self, -1, u'开始爬取限速', pos=(20, 280), size=(100, 25))
        self.btn_cancel2 = wx.Button(self, -1, u'取消爬取限速', pos=(140, 280), size=(100, 25))

        self.btn_cancel2.Disable()

        # 控件事件
        #self.gd_key.Bind(wx.EVT_TEXT, self.EvtText)

        # 鼠标事件
        #self.Bind(wx.EVT_BUTTON, self.OnClose, self.btn_close)
        self.btn_start.Bind(wx.EVT_LEFT_DOWN, self.OnStartDown)
        self.Bind(wx.EVT_BUTTON, self.OnPauseDown, self.btn_pause)
        self.Bind(wx.EVT_COMBOBOX, self.OnProvinceChoice, self.ch1)
        self.Bind(wx.EVT_COMBOBOX, self.OnCityChoice, self.ch2)

        self.btn_start2.Bind(wx.EVT_LEFT_DOWN, self.StartReptileRoad)
        self.btn_cancel2.Bind(wx.EVT_LEFT_DOWN, self.CancelReptileRoad)
        # 暂停按钮启动时不可点击
        self.btn_pause.Disable()
        # 创建异步执行爬取数据的线程
        self.t1 = threading.Thread(target=self.startWork, args=(self.gd_key.GetValue(),))
        self.t2 = threading.Thread(target=self.startWork2, args=(self.gd_key.GetValue(),))

        # 系统事件
        self.Bind(wx.EVT_CLOSE, self.OnClose)
        self.Bind(wx.EVT_SIZE, self.On_size)
        #self.Bind(wx.EVT_PAINT, self.On_paint)
        #self.Bind(wx.EVT_ERASE_BACKGROUND, lambda event: None)

    def EvtText(self, evt):
        '''输入框事件函数'''
        obj = evt.GetEventObject()
        objName = obj.GetName()
        text = evt.GetString()

        if objName == 'TC01':
            print(text)
        elif objName == 'TC02':
            print(text)

    def On_size(self, evt):
        # 改变窗口大小事件函数
        self.Refresh()
        evt.Skip()

    workbook = Workbook()
    file1 =''
    iswriting=False   # 是否正在开始写xls
    ispause = False   # 是否暂停
    isreptiling = False  #是否在爬取中
    isrunsched =False  #是否开启了调度
    ispause2=False     # 是否开启爬取限速

    def OnClose(self, evt):
        # 关闭窗口事件函数
        dlg = wx.MessageDialog(None, u'确定要关闭本窗口？', u'操作提示', wx.YES_NO | wx.ICON_QUESTION)
        if(dlg.ShowModal() == wx.ID_YES):
            if self.iswriting:
                self.workbook.save(self.file1)
            if self.isrunsched:
                self.scheduler.shutdown()
            self.Destroy()


    def OnStartDown(self, evt):
        if self.loat.GetValue() =='' and (self.city == '' or self.city == '-1'):
            self.area.AppendText('[warn]请设置经纬度矩形区域，或选择城市！！！\n')
            return
        key = self.gd_key.GetValue()
        if key == '':
            key = '0b1804994cd63974f873a29a269d65e7'
            self.area.AppendText('[warn]请填写高德web服务key！！！\n')
            return
        self.loat.Disable()
        self.gd_key.Disable()
        self.btn_start.Disable()
        self.btn_pause.Enable()
        self.ch1.Disable()
        self.ch2.Disable()

        if self.ispause:
            if self.preCity == '-1':
                self.preCity = self.city
            if self.preCity != self.city:
                # 暂停后又开始时切换了城市，原来已爬取的城市立即保存，开始爬取新城市
                if self.iswriting:
                    self.workbook.save(self.file1)

            self.area.AppendText('[info]恢复爬取\n')
            self.ispause=False
            self.scheduler.resume()
        else:
            self.area.AppendText('[info]key='+key+'\n')
            self.t1 = threading.Thread(target=self.startWork, args=(key,))
            self.t1.setDaemon(True)  # 设置为守护线程
            self.t1.start()

    def OnPauseDown(self, evt):
        self.loat.Enable()
        self.gd_key.Enable()
        self.btn_start.Enable()
        self.btn_pause.Disable()
        self.ch1.Enable()
        self.ch2.Enable()
        self.pauseWork()

    def OnProvinceChoice(self,evt):
        province = evt.GetString()
        citys = city.province_city[province]
        self.ch2.Clear()
        for key in citys:
            self.ch2.Append(key)

    def OnCityChoice(self,evt):
        self.city = evt.GetString()
        print(self.city)
        self.cityTip.SetLabelText('当前城市：'+self.city)


    def LocaDiv2(self, ploy):
        list = []
        p0 = float(ploy.split(',')[0])
        p1 = float(ploy.split(',')[1])
        p2 = float(ploy.split(',')[2])
        p3 = float(ploy.split(',')[3])
        len1 = int((p2 - p0 + 0.0001) / 0.05)
        len2 = int((p3 - p1 + 0.0001) / 0.04)
        for i in range(0, len1):
            for j in range(0, len2):
                a = round(p0 + round(0.05 * i, 2), 6)
                b = round(p1 + round(0.04 * j, 2), 6)
                c = round(a+round(0.05 * 1, 2), 6)
                d = round(b+round(0.04 * 1, 2), 6)
                pos = str(a)+','+str(b)+';'+str(c)+','+str(d)
                list.append(pos)
        return list

    def reptileMap(self, key):
        print('key='+key)
        print('[info]开始爬取数据:'+self.city)
        self.area.Clear()
        self.area.AppendText('[info]使用key'+key)
        self.area.AppendText('[info]开始爬取数据...\n')
        startTime = time.time()

        lotc=''
        if self.loat.GetValue() !='':
            lotc = self.loat.GetValue()
        else:
            lotc = city.city_pos[self.city]
        print(lotc)
        locs = self.LocaDiv2(lotc)
        date = time.strftime("%Y%m%d-%H")

        dirs = os.path.abspath('.')+'\\'+self.city + '\\' + time.strftime("%Y%m%d")
        if self.loat.GetValue() !='':
            dirs = os.path.abspath('.')+'\\'+self.loat.GetValue() + '\\' + time.strftime("%Y%m%d")
        # 创建文件夹
        if not os.path.exists(dirs):
            os.makedirs(dirs)
        # 删除旧文件
        self.file1 = dirs+'\\'+  date +'.xlsx'
        if os.path.exists(self.file1):
            os.remove(self.file1)

        dttime = time.strftime("%Y-%m-%d %H:%M:%S")
        count = 1
        self.workbook = Workbook()
        sheet1 = self.workbook.create_sheet(self.city,0)
        keys1 = ['angle', 'direction', 'lcodes', 'name', 'polyline', 'speed', 'status', 'description', 'evaluation', 'datetime']
        for i in range(0, len(keys1)):
            sheet1.cell(row=1, column=i+1).value = keys1[i]  # 写入表头

        self.iswriting =True
        # if self.isreptiling != True:
        #     self.workbook.save(self.file1)
        #     return

        self.isreptiling = True

        for loc in locs:
            pa = {
                'key': str(key),
                # 'level': 6,                   # 道路等级为6，即返回的道路路况等级最小到无名道路这一级别
                'rectangle': str(loc),          # 矩形区域
                'extensions': 'all'
                # 'output': 'JSON'
            }
            print('[info]探测区块：'+loc)
            self.area.AppendText('[info]探测区块：'+loc+'\n')
            obj = '{}'
            while True:
                try:
                    obj = requests.get('http://restapi.amap.com/v3/traffic/status/rectangle?', params=pa, timeout=30)
                    break
                except requests.exceptions.ConnectionError:
                    print('[ERROR]ConnectionError -- will retry connect')
                    self.area.AppendText('[ERROR]ConnectionError -- will retry connect\n')
                    time.sleep(3)

            data = obj.json()
            if data['status'] == '0':
                print('[info]'+str(data))
                print('[warn]请求参数错误')
                self.area.AppendText('[warn]'+str(data)+'\n')
                continue


            # print(data)

            description = data['trafficinfo']['description'] if 'description' in data['trafficinfo'] else ''
            evaluation = data['trafficinfo']['evaluation'] if 'evaluation' in data['trafficinfo'] else ''
            evaluation = str(evaluation)

            for road in data['trafficinfo']['roads']:
                count = count+1

                rangle = road['angle'] if 'angle' in road else '0'
                rdirection = road['direction'] if 'direction' in road else ''
                rlcodes = road['lcodes'] if 'lcodes' in road else ''
                rname = road['name'] if 'name' in road else ''
                rpolyline = road['polyline'] if 'polyline' in road else ''
                rspeed = road['speed'] if 'speed' in road else '0'
                rstatus = road['status'] if 'status' in road else '0'

                sheet1.cell(row=count, column=1).value = int(rangle)
                sheet1.cell(row=count, column=2).value = rdirection
                sheet1.cell(row=count, column=3).value = rlcodes
                sheet1.cell(row=count, column=4).value = rname
                sheet1.cell(row=count, column=5).value = rpolyline
                sheet1.cell(row=count, column=6).value = int(rspeed)
                sheet1.cell(row=count, column=7).value = int(rstatus)
                sheet1.cell(row=count, column=8).value = description
                sheet1.cell(row=count, column=9).value = evaluation
                sheet1.cell(row=count, column=10).value = dttime

            time.sleep(1)    # 间隔1s执行一次分块请求，避免并发度高被限制
        self.workbook.save(self.file1)
        endTime = time.time()
        print('[info]数据爬取完毕，用时%.2f秒' % (endTime-startTime))
        print('[info]数据存储路径：'+self.file1)
        self.area.AppendText('[info]数据爬取完毕，用时%.2f秒' % (endTime-startTime) +'\n')
        self.area.AppendText('[info]数据存储路径：'+self.file1 +'\n')
        self.isreptiling = False
        self.iswriting= False

    scheduler = BlockingScheduler()
    def startWork(self, key):
        # 开始首次爬取，首次爬取过程中不会暂停，其他时候可以暂停
        self.reptileMap(key)
        self.isrunsched = True
        # 周一到周日,每小时执行一次   每5秒second='*/5' hour='0-23'
        trigger = CronTrigger(day_of_week='0-6', hour='0-23')
        self.scheduler.add_job(self.reptileMap, trigger, args=(key,))
        self.scheduler.start()

    def pauseWork(self):
        self.area.AppendText('[info]暂停爬取数据...\n')
        self.ispause=True
        self.scheduler.pause()


    scheduler2 = BlockingScheduler()
    def StartReptileRoad(self, evt):
        key = self.gd_key.GetValue()
        if key == '':
            key = '0b1804994cd63974f873a29a269d65e7_1'
            self.area.AppendText('[warn]请填写高德web服务key！！！\n')
            return
        self.btn_start2.Disable()
        self.btn_cancel2.Enable()

        self.reptileRoad(key)
        if self.ispause2:
            self.ispause2=False
            self.scheduler2.resume()
        else:
            self.t2 = threading.Thread(target=self.startWork2, args=(key,))
            self.t2.setDaemon(True)  # 设置为守护线程
            self.t2.start()

    def startWork2(self,key):
        # 定时每天 01:00:30秒执行任务
        trigger2 = CronTrigger(day_of_week='0-6', hour = 1,minute = 0,second = 30 )
        self.scheduler2.add_job(self.reptileRoad, trigger2, args=(key,))
        self.scheduler2.start()


    def CancelReptileRoad(self, evt):
        self.btn_start2.Enable()
        self.btn_cancel2.Disable()
        self.ispause2=True
        self.scheduler2.pause()


    def reptileRoad(self,key):
        # 获取道路文件路径
        for root, dirs, files in os.walk(os.path.abspath('.')):
           dir=time.strftime("%Y%m%d")
           if root.endswith(dir) and len(files)>0:
               filePath = root+"\\"+files[0]
               print('file:', filePath)
               self.area.AppendText('[info]文件：'+filePath +'，开始爬取道路限速\n')
               wb = load_workbook(filePath)
               sheet = wb.active
               rnum=sheet.max_row
               cnum=sheet.max_column
               sheet.cell(row=1, column=11).value = 'roadlevel'
               sheet.cell(row=1, column=12).value = 'maxspeed'
               for r in range(2, rnum):
                   polyline =sheet.cell(row=r, column=5).value
                   s1 = polyline.replace(';', '|') # 点参数
                   remove_digits = str.maketrans('', '', digits)
                   s2 = polyline.translate(remove_digits).replace('.,.', '1').replace(';', ',')
                   arr = s2.split(",")  #方向速度参数
                   s3 = ""  # 时间戳参数
                   t = time.time()
                   for i in range(0,len(arr)):
                       tt=int(t)+i
                       s3 = s3+str(tt)+','
                   s3 = s3[:len(s3)-1]
                   param={
                      'key': str(key),
                      'extensions': 'all',
                      'carid': 'ts001',
                      'locations': s1,
                      'direction': s2,
                      'speed': s2,
                      'time': s3
                   }
                   obj = '{}'
                   while True:
                       try:
                           obj =requests.get('https://restapi.amap.com/v3/autograsp?', params=param, timeout=30)
                           break
                       except requests.exceptions.ConnectionError:
                           print('[ERROR]ConnectionError2 -- will retry connect')
                           self.area.AppendText('[ERROR]ConnectionError2 -- will retry connect\n')
                           time.sleep(1)

                   data = obj.json()
                   if data['status'] == '0':
                       print('[info]'+str(data))
                       print('[warn]请求参数错误',str(param))
                       continue
                   if data['status'] == '1':
                       for road in data['roads']:
                           maxspeed= road.get("maxspeed")
                           roadlevel= road.get("roadlevel")
                           if maxspeed is None or maxspeed =='' or maxspeed == "-1":
                               continue
                           else:
                               sheet.cell(row=r, column=11).value = int(roadlevel)
                               sheet.cell(row=r, column=12).value = int(maxspeed)
                               break
                   # if r>10:
                   #     break
               wb.save(filePath)
               print(filePath, "爬取道路限速和道路等级成功")
               self.area.AppendText('[info]爬取道路限速成功，数据存储路径：'+filePath +'\n')


class mainApp(wx.App):
    def OnInit(self):
        self.SetAppName(APP_TITLE)
        self.Frame = mainFrame(None)
        self.Frame.Show()
        return True

if __name__ == "__main__":
    app = mainApp()
    app.MainLoop()
